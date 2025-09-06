import csv
import openpyxl
from io import StringIO, BytesIO
from django.db.models import Q
from django.db import transaction
from django.db.models import Count, QuerySet
from django.utils import timezone
from apps.study.models import UserTestAttempt
from apps.users.models import UserProfile
from rest_framework.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _

from apps.learning.models import (
    Question,
    TestType,
    LearningSection,
    LearningSubSection,
    Skill,
    MediaFile,
    Article,
)


def get_filtered_test_attempts(filters: dict):
    """
    Retrieves and filters UserTestAttempt queryset based on provided filters.
    This is the single source of truth for querying export data.
    """
    datetime_from = filters.get("datetime_from")
    datetime_to = filters.get("datetime_to")

    # Start with a base queryset.
    queryset = UserTestAttempt.objects.filter(status=UserTestAttempt.Status.COMPLETED)

    # Apply date filters if they exist.
    # Note: The view now passes date strings, so we can filter directly.
    if datetime_from and datetime_to:
        queryset = queryset.filter(start_time__range=(datetime_from, datetime_to))
    elif datetime_from:
        queryset = queryset.filter(start_time__gte=datetime_from)
    elif datetime_to:
        queryset = queryset.filter(start_time__lte=datetime_to)

    # --- OPTIMIZATION & ENRICHMENT ---
    # Eagerly load related data to prevent N+1 queries in the loop.
    # We need user details and test definition details.
    queryset = queryset.select_related("user", "test_definition")

    # Use annotation to efficiently count answered questions for each attempt.
    # This prevents a separate DB query for every row in the export.
    queryset = queryset.annotate(answered_question_count_agg=Count("question_attempts"))

    # Add any other future filters here from the 'filters' dict
    # e.g., if filters.get('user_id'): queryset = queryset.filter(...)

    return queryset.order_by("-start_time")


def generate_export_file_content(queryset, export_format: str):
    """
    Generates file content (CSV or XLSX) from a queryset.
    This is the single source of truth for file generation logic.
    Provides a rich, meaningful set of columns for analysis.
    """
    # --- NEW, MORE INFORMATIVE HEADERS ---
    headers = [
        "Attempt ID",
        "User ID",
        "Username",
        "User Full Name",
        "User Email",
        "Attempt Type",
        "Test Name",
        "Test Definition Type",
        "Status",
        "Start Time (UTC)",
        "End Time (UTC)",
        "Duration (Minutes)",
        "Total Questions in Test",
        "Questions Answered",
        "Overall Score (%)",
        "Verbal Score (%)",
        "Quantitative Score (%)",
    ]
    filename = f"qader_test_attempts_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{export_format}"
    content_type = None
    file_content = None

    # Use iterator() to handle large querysets efficiently by processing
    # records in chunks, reducing memory usage.
    if export_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for attempt in queryset.iterator():
            # Handle cases where related objects might be null
            test_def_name = (
                attempt.test_definition.name
                if attempt.test_definition
                else "N/A (Traditional Practice)"
            )
            test_def_type = (
                attempt.test_definition.get_test_type_display()
                if attempt.test_definition
                else "N/A"
            )
            duration_minutes = (
                round(attempt.duration_seconds / 60, 2)
                if attempt.duration_seconds is not None
                else None
            )

            writer.writerow(
                [
                    attempt.id,
                    attempt.user.id,
                    attempt.user.username,
                    attempt.user.get_full_name(),
                    attempt.user.email,
                    attempt.get_attempt_type_display(),
                    test_def_name,
                    test_def_type,
                    attempt.get_status_display(),
                    attempt.start_time.strftime("%Y-%m-%d %H:%M:%S"),
                    (
                        attempt.end_time.strftime("%Y-%m-%d %H:%M:%S")
                        if attempt.end_time
                        else None
                    ),
                    duration_minutes,
                    attempt.num_questions,
                    attempt.answered_question_count_agg,  # Use the efficient annotated value
                    attempt.score_percentage,
                    attempt.score_verbal,
                    attempt.score_quantitative,
                ]
            )
        file_content = output.getvalue().encode("utf-8")
        content_type = "text/csv"

    elif export_format == "xlsx":
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Test Attempts"
        sheet.append(headers)
        for attempt in queryset.iterator():
            # Handle cases where related objects might be null
            test_def_name = (
                attempt.test_definition.name
                if attempt.test_definition
                else "N/A (Traditional Practice)"
            )
            test_def_type = (
                attempt.test_definition.get_test_type_display()
                if attempt.test_definition
                else "N/A"
            )
            duration_minutes = (
                round(attempt.duration_seconds / 60, 2)
                if attempt.duration_seconds is not None
                else None
            )

            # Excel doesn't handle timezone-aware datetimes well, so we remove tzinfo
            start_time_naive = attempt.start_time.replace(tzinfo=None)
            end_time_naive = (
                attempt.end_time.replace(tzinfo=None) if attempt.end_time else None
            )

            sheet.append(
                [
                    attempt.id,
                    attempt.user.id,
                    attempt.user.username,
                    attempt.user.get_full_name(),
                    attempt.user.email,
                    attempt.get_attempt_type_display(),
                    test_def_name,
                    test_def_type,
                    attempt.get_status_display(),
                    start_time_naive,
                    end_time_naive,
                    duration_minutes,
                    attempt.num_questions,
                    attempt.answered_question_count_agg,  # Use the efficient annotated value
                    attempt.score_percentage,
                    attempt.score_verbal,
                    attempt.score_quantitative,
                ]
            )
        workbook.save(output)
        file_content = output.getvalue()
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return file_content, content_type, filename


# --- NEW SERVICE FUNCTIONS FOR USER EXPORT ---


def get_filtered_users(filters: dict):
    """
    Retrieves and filters UserProfile queryset based on provided filters.
    This is the single source of truth for querying user export data.
    """
    queryset = UserProfile.objects.select_related(
        "user", "assigned_mentor__user", "referred_by"
    ).all()

    # --- APPLY FILTERS FROM THE 'filters' DICTIONARY ---
    roles_to_filter = filters.get("role")

    # Check if a list of roles was provided and it's not empty
    if roles_to_filter:
        queryset = queryset.filter(role__in=roles_to_filter)

    # You can add more filters here in the future
    # account_type = filters.get('account_type')
    # if account_type:
    #     queryset = queryset.filter(account_type=account_type)

    return queryset.order_by("-user__date_joined")


def generate_user_export_file_content(queryset, export_format: str):
    """
    Generates file content (CSV or XLSX) for User data from a queryset.
    """
    headers = [
        "User ID",
        "Username",
        "Full Name",
        "Preferred Name",
        "Email",
        "Role",
        "Account Type",
        "Is Active",
        "Is Subscribed",
        "Subscription Expires At (UTC)",
        "Date Joined (UTC)",
        "Last Login (UTC)",
        "Gender",
        "Grade",
        "Taken Qiyas Before",
        "Points",
        "Current Streak (Days)",
        "Longest Streak (Days)",
        "Verbal Level (%)",
        "Quantitative Level (%)",
        "Referral Code",
        "Referred By (Username)",
        "Assigned Mentor",
    ]
    filename = f"qader_users_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{export_format}"
    content_type = None
    file_content = None

    if export_format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for profile in queryset.iterator():
            writer.writerow(
                [
                    profile.user_id,
                    profile.user.username,
                    profile.full_name,
                    profile.preferred_name,
                    profile.user.email,
                    profile.get_role_display(),
                    profile.get_account_type_display(),
                    profile.user.is_active,
                    profile.is_subscribed,
                    (
                        profile.subscription_expires_at.strftime("%Y-%m-%d %H:%M:%S")
                        if profile.subscription_expires_at
                        else None
                    ),
                    profile.user.date_joined.strftime("%Y-%m-%d %H:%M:%S"),
                    (
                        profile.user.last_login.strftime("%Y-%m-%d %H:%M:%S")
                        if profile.user.last_login
                        else None
                    ),
                    profile.get_gender_display(),
                    profile.get_grade_display(),
                    profile.has_taken_qiyas_before,
                    profile.points,
                    profile.current_streak_days,
                    profile.longest_streak_days,
                    profile.current_level_verbal,
                    profile.current_level_quantitative,
                    profile.referral_code,
                    profile.referred_by.username if profile.referred_by else None,
                    (
                        profile.assigned_mentor.user.username
                        if profile.assigned_mentor
                        else None
                    ),
                ]
            )
        file_content = output.getvalue().encode("utf-8")
        content_type = "text/csv"

    elif export_format == "xlsx":
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(headers)
        for profile in queryset.iterator():
            sheet.append(
                [
                    profile.user_id,
                    profile.user.username,
                    profile.full_name,
                    profile.preferred_name,
                    profile.user.email,
                    profile.get_role_display(),
                    profile.get_account_type_display(),
                    profile.user.is_active,
                    profile.is_subscribed,
                    (
                        profile.subscription_expires_at.replace(tzinfo=None)
                        if profile.subscription_expires_at
                        else None
                    ),
                    profile.user.date_joined.replace(tzinfo=None),
                    (
                        profile.user.last_login.replace(tzinfo=None)
                        if profile.user.last_login
                        else None
                    ),
                    profile.get_gender_display(),
                    profile.get_grade_display(),
                    profile.has_taken_qiyas_before,
                    profile.points,
                    profile.current_streak_days,
                    profile.longest_streak_days,
                    profile.current_level_verbal,
                    profile.current_level_quantitative,
                    profile.referral_code,
                    profile.referred_by.username if profile.referred_by else None,
                    (
                        profile.assigned_mentor.user.username
                        if profile.assigned_mentor
                        else None
                    ),
                ]
            )
        workbook.save(output)
        file_content = output.getvalue()
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return file_content, content_type, filename


# --- NEW SERVICE FUNCTIONS FOR QUESTION IMPORT/EXPORT ---


def get_filtered_questions(filters: dict) -> QuerySet[Question]:
    """
    Applies filters to the Question queryset for export, mirroring AdminQuestionViewSet.
    """
    queryset = (
        Question.objects.select_related(
            "subsection__section__test_type", "media_content", "article"
        )
        .prefetch_related("skills")
        .all()
    )

    # Apply filters from the 'filters' dictionary passed by the view
    if filters.get("subsection__section__test_type__id"):
        queryset = queryset.filter(
            subsection__section__test_type__id=filters[
                "subsection__section__test_type__id"
            ]
        )
    if filters.get("subsection__section__id"):
        queryset = queryset.filter(
            subsection__section__id=filters["subsection__section__id"]
        )
    # ... Add any other filters from AdminQuestionViewSet as needed ...
    if filters.get("is_active"):
        is_active_bool = str(filters["is_active"]).lower() in ["true", "1"]
        queryset = queryset.filter(is_active=is_active_bool)

    return queryset.order_by("id")


ARABIC_QUESTION_HEADERS = [
    "معرف السؤال",
    "نص السؤال",
    "نشط",
    "خيار أ",
    "خيار ب",
    "خيار ج",
    "خيار د",
    "الإجابة الصحيحة",
    "الشرح",
    "تلميح",
    "ملخص الحل",
    "مستوى الصعوبة",
    "نوع الاختبار",
    "القسم الرئيسي",
    "القسم الفرعي",
    "المهارات",
    "عنوان الوسائط",
    "عنوان المقال",
]

ARABIC_DIFFICULTY_TO_INT = {
    str(_("Very Easy")): 1,
    str(_("Easy")): 2,
    str(_("Medium")): 3,
    str(_("Hard")): 4,
    str(_("Very Hard")): 5,
    "1": 1,
    "2": 2,
    "3": 3,
    "4": 4,
    "5": 5,
    1: 1,
    2: 2,
    3: 3,
    4: 4,
    5: 5,
}

LATIN_TO_ARABIC_CHOICE = {
    "A": "أ",
    "B": "ب",
    "C": "ج",
    "D": "د",
}

ARABIC_TO_LATIN_CHOICE = {
    "أ": "A",
    "ب": "B",
    "ج": "C",
    "د": "D",
}


def generate_question_export_file_content(
    queryset: QuerySet[Question], export_format: str = "xlsx"
):
    """
    Generates a user-friendly Excel (XLSX) file for questions with Arabic headers.
    """
    headers = ARABIC_QUESTION_HEADERS  # Use Arabic headers
    filename = (
        f"qader_questions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{export_format}"
    )

    if export_format != "xlsx":
        raise NotImplementedError("Only XLSX format is supported for question export.")

    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "الأسئلة"  # Sheet title in Arabic
    sheet.append(headers)

    # Set sheet direction to Right-to-Left for better Arabic readability in Excel
    sheet.sheet_view.rightToLeft = True

    for q in queryset.iterator(chunk_size=2000):
        skill_names = ", ".join(skill.name for skill in q.skills.all())

        correct_answer_arabic = LATIN_TO_ARABIC_CHOICE.get(
            q.correct_answer, q.correct_answer
        )

        sheet.append(
            [
                q.id,
                q.question_text,
                1 if q.is_active else 0,  # MODIFIED: Use 1/0 for Is Active
                q.option_a,
                q.option_b,
                q.option_c,
                q.option_d,
                correct_answer_arabic,
                q.explanation,
                q.hint,
                q.solution_method_summary,
                q.get_difficulty_display(),  # Still export the human-readable Arabic name
                getattr(q.subsection.section.test_type, "name", ""),
                getattr(q.subsection.section, "name", ""),
                getattr(q.subsection, "name", ""),
                skill_names,
                getattr(q.media_content, "title", ""),
                getattr(q.article, "title", ""),
            ]
        )

    workbook.save(output)
    file_content = output.getvalue()
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return file_content, content_type, filename


@transaction.atomic
def process_question_import_file(file_obj, update_strategy: str):
    """
    Processes an uploaded Excel file with Arabic headers to create or update questions.
    """
    try:
        workbook = openpyxl.load_workbook(file_obj)
        sheet = workbook.active
    except Exception as e:
        raise ValidationError(f"Could not read the Excel file. Error: {str(e)}")

    headers = [cell.value for cell in sheet[1]]
    expected_headers = ARABIC_QUESTION_HEADERS  # Check against Arabic headers

    if headers != expected_headers:
        raise ValidationError(
            {
                "error": "Invalid file headers.",
                "expected_headers": expected_headers,
                "found_headers": headers,
            }
        )

    errors = []
    created_count = 0
    updated_count = 0

    # Create a map from header name to its index for robust data access
    header_map = {header: i for i, header in enumerate(headers)}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue  # Skip empty rows

        # Helper to get data by Arabic header name
        def get_data(header_name):
            return row_values[header_map[header_name]]

        try:
            # --- MODIFIED: Use get_or_create with Arabic headers ---
            test_type_name = get_data("نوع الاختبار")
            if not test_type_name:
                raise ValueError("'نوع الاختبار' is required.")
            test_type, _ = TestType.objects.get_or_create(name=test_type_name)

            section_name = get_data("القسم الرئيسي")
            if not section_name:
                raise ValueError("'القسم الرئيسي' is required.")
            section, _ = LearningSection.objects.get_or_create(
                name=section_name, test_type=test_type
            )

            subsection_name = get_data("القسم الفرعي")
            if not subsection_name:
                raise ValueError("'القسم الفرعي' is required.")
            subsection, _ = LearningSubSection.objects.get_or_create(
                name=subsection_name, section=section
            )

            skills_to_set = []
            skill_names_str = get_data("المهارات")
            if skill_names_str:
                skill_names_list = [
                    name.strip()
                    for name in str(skill_names_str).split(",")
                    if name.strip()
                ]
                for skill_name in skill_names_list:
                    skill, _ = Skill.objects.get_or_create(
                        name=skill_name, section=section, subsection=subsection
                    )
                    skills_to_set.append(skill)

            media_title = get_data("عنوان الوسائط")
            media_content = (
                MediaFile.objects.get(title=media_title) if media_title else None
            )

            article_title = get_data("عنوان المقال")
            article = (
                Article.objects.get(title=article_title) if article_title else None
            )

            # --- Data Cleaning with Arabic/Numeric values ---
            difficulty_input = get_data("مستوى الصعوبة")
            difficulty_value = ARABIC_DIFFICULTY_TO_INT.get(difficulty_input)
            if difficulty_value is None:
                raise ValueError(
                    f"قيمة مستوى الصعوبة غير صالحة: '{difficulty_input}'. يجب أن تكون أحد الخيارات: {list(ARABIC_DIFFICULTY_TO_INT.keys())}"
                )

            is_active = get_data("نشط") == 1  # Check if value is 1

            correct_answer_input = get_data("الإجابة الصحيحة")
            correct_answer_latin = ARABIC_TO_LATIN_CHOICE.get(correct_answer_input)

            if correct_answer_latin is None:
                # Also check if they entered Latin letters directly, for convenience
                if str(correct_answer_input).upper() in LATIN_TO_ARABIC_CHOICE:
                    correct_answer_latin = str(correct_answer_input).upper()
                else:
                    raise ValueError(
                        f"قيمة الإجابة الصحيحة غير صالحة: '{correct_answer_input}'. يجب أن تكون أحد الخيارات: أ, ب, ج, د"
                    )

            # --- Create or Update Logic (Unchanged) ---
            question_id = get_data("معرف السؤال")

            question = (
                Question.objects.filter(id=question_id).first() if question_id else None
            )
            if question and update_strategy == "SKIP":
                continue
            if not question:
                question = Question()
                is_update = False
            else:
                is_update = True

            question.question_text = get_data("نص السؤال")
            question.is_active = is_active
            question.option_a, question.option_b = get_data("خيار أ"), get_data(
                "خيار ب"
            )
            question.option_c, question.option_d = get_data("خيار ج"), get_data(
                "خيار د"
            )
            question.correct_answer = correct_answer_latin
            question.difficulty = difficulty_value
            question.explanation, question.hint = get_data("الشرح"), get_data("تلميح")
            question.solution_method_summary = get_data("ملخص الحل")
            question.subsection = subsection
            question.media_content, question.article = media_content, article

            question.full_clean()
            question.save()

            if skills_to_set:
                question.skills.set(skills_to_set)
            else:
                question.skills.clear()

            if is_update:
                updated_count += 1
            else:
                created_count += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    if errors:
        raise ValidationError(
            {"detail": "فشل الاستيراد بسبب وجود أخطاء.", "errors": errors}
        )

    return {"created": created_count, "updated": updated_count}
